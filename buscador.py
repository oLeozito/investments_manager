import requests
from bs4 import BeautifulSoup
import json
import os
from datetime import datetime

# Import pra manipular excel
import openpyxl
from openpyxl.styles import Alignment, Font


class FundoImobiliario:
    def __init__(self, codigo, dy_12m, dy_percent, preco_atual, segmento, tipo, val_patr, vacancia, qtdcotis, qtdcotas, cnpj, preco_teto, pvp, liquidez):
        self.codigo = codigo
        self.dy_12m = dy_12m
        self.dy_percent = dy_percent
        self.preco_atual = preco_atual
        self.segmento = segmento
        self.tipo = tipo
        self.val_patr = val_patr
        self.vacancia = vacancia
        self.qtdcotis = qtdcotis
        self.qtdcotas = qtdcotas
        self.cnpj = cnpj
        self.pvp = pvp
        self.liquidez = liquidez
        self.preco_teto = preco_teto

    def to_dict(self):
        return self.__dict__

    @staticmethod
    def from_dict(dados):
        return FundoImobiliario(**dados)

class Data:
    def __init__(self, siglas=None, fundos=None, ultima_atualizacao=None):
        self.siglas = siglas if siglas else []
        self.fundos = fundos if fundos else {}
        self.ultima_atualizacao = ultima_atualizacao

    def salvar(self, arquivo="data.json"):
        dados = {
            "siglas": self.siglas,
            "fundos": self.fundos,
            "ultima_atualizacao": self.ultima_atualizacao
        }
        with open(arquivo, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)

    @staticmethod
    def carregar(arquivo="data.json"):
        if os.path.exists(arquivo):
            with open(arquivo, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return Data(
                    siglas=dados.get("siglas", []),
                    fundos=dados.get("fundos", {}),
                    ultima_atualizacao=dados.get("ultima_atualizacao")
                )
        else:
            return Data()

def parse_brl_number(texto):
    if not texto or not isinstance(texto, str):
        return None
    texto = texto.replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(texto)
    except:
        return None

def parse_monetary_value(texto):
    if not texto:
        return None
    texto = texto.replace("R$", "").strip()

    if "Milhões" in texto or "Milhão" in texto:
        texto = texto.replace("Milhões", "").replace("Milhão", "").strip()
        return parse_brl_number(texto) * 1_000_000
    elif "Bilhões" in texto or "Bilhão" in texto:
        texto = texto.replace("Bilhões", "").replace("Bilhão", "").strip()
        return parse_brl_number(texto) * 1_000_000_000
    else:
        return parse_brl_number(texto)

def parse_liquidez(texto):
    if not texto:
        return None
    texto = texto.replace("R$", "").strip()

    multiplicador = 1
    if "K" in texto:
        multiplicador = 1_000
        texto = texto.replace("K", "").strip()
    elif "M" in texto:
        multiplicador = 1_000_000
        texto = texto.replace("M", "").strip()

    valor = parse_brl_number(texto)
    return valor * multiplicador if valor is not None else None

def parse_percent(texto):
    if not texto:
        return None
    texto = texto.replace("%", "").replace(",", ".").strip()
    try:
        return float(texto)
    except:
        return None

def parse_int(texto):
    if not texto:
        return None
    texto = texto.replace(".", "").strip()
    try:
        return int(texto)
    except:
        return None

def obter_fiis():
    letras = [chr(i) for i in range(ord('A'), ord('Z')+1)]
    siglas = []

    url_fiis = "https://www.fundsexplorer.com.br/funds"
    headers = {'User-Agent': 'Mozilla/5.0'}

    resposta = requests.get(url_fiis, headers=headers)
    if resposta.status_code != 200:
        print(f"Erro ao acessar o site.")

    html_fiis = BeautifulSoup(resposta.text, 'html.parser')
    listar = html_fiis.find('div', class_='tickerFilter__results')

    for letra in letras:
        filtrar = listar.find('section', id=f"letter-id-{letra}")
        buscar = filtrar.find_all('div', attrs={'data-element': 'ticker-box-title'})
        for item in buscar:
            siglas.append(item.text.strip())

    print(f"\nBusca de {len(siglas)} siglas completa.\n")
    return siglas

def obter_infos(fii):
    url = f'https://investidor10.com.br/fiis/{fii.lower()}/'
    headers = {'User-Agent': 'Mozilla/5.0'}

    resposta = requests.get(url, headers=headers)
    if resposta.status_code != 200:
        print(f"Erro ao acessar o site para {fii}")
        return None

    soup = BeautifulSoup(resposta.text, 'html.parser')

    spans_dy = soup.find_all('span', class_='content--info--item--value amount')
    valores_rs = [parse_brl_number(span.text) for span in spans_dy if 'R$' in span.text]
    val_dy_12m = valores_rs[3] if len(valores_rs) >= 4 else None

    preco_cota = None
    for span in soup.find_all('span', class_='value'):
        if 'R$' in span.text:
            preco_cota = parse_brl_number(span.text)
            if preco_cota:
                break

    def get_card_value(classe):
        card = soup.find('div', class_=f'_card {classe}')
        if card:
            corpo = card.find('div', class_='_card-body')
            if corpo:
                span = corpo.find('span')
                if span:
                    return span.text.strip()
        return None

    dy12 = parse_percent(get_card_value("dy"))
    pvp = parse_brl_number(get_card_value("vp"))
    liquidez = parse_liquidez(get_card_value("val"))

    valores = {}
    tabela = soup.find('div', id="table-indicators")
    if tabela:
        celulas = tabela.find_all('div', class_="cell")
        for cell in celulas:
            nome = cell.find('span', class_="d-flex justify-content-between align-items-center name")
            value = cell.find('div', class_="value")
            if nome and value:
                valores[nome.text.strip()] = value.text.strip()

    cnpj       = valores.get("CNPJ")
    segmento   = valores.get("SEGMENTO")
    tipo       = valores.get("TIPO DE FUNDO")
    vacancia   = parse_percent(valores.get("VACÂNCIA"))
    qtd_cotis  = parse_int(valores.get("NUMERO DE COTISTAS"))
    qtd_cotas  = parse_int(valores.get("COTAS EMITIDAS"))
    val_patri  = parse_monetary_value(valores.get("VALOR PATRIMONIAL"))

    preco_teto = round(val_dy_12m / 0.12, 2) if val_dy_12m else None

    return FundoImobiliario(
        fii, val_dy_12m, dy12, preco_cota, segmento, tipo,
        val_patri, vacancia, qtd_cotis, qtd_cotas, cnpj,
        preco_teto, pvp, liquidez
    )

def limpar_tabela(pagina):
    for row in pagina.iter_rows(min_row = 3, max_row = 100, min_col=1, max_col=12):
        for cell in row:
            cell.value = None
    print("\nTabela Limpa")

def preencher_tabela(fundos,pagina):
    count = 0
    chaves = list(fundos.keys())
    for row in pagina.iter_rows(min_row =3, max_row =len(fundos.keys()),min_col=1,max_col=11):
        flag = 0
        for cell in row:
                if flag == 0:
                    cell.value = fundos[chaves[count]].codigo
                    cell.hyperlink = f"https://investidor10.com.br/fiis/{(fundos[chaves[count]].codigo).lower()}/"
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                    flag+=1
                elif flag == 1:
                    cell.value = fundos[chaves[count]].segmento
                    flag+=1
                elif flag == 2:
                    cell.value = fundos[chaves[count]].tipo
                    flag+=1
                elif flag == 3:
                    cell.value = fundos[chaves[count]].preco_atual
                    flag+=1
                elif flag == 4:
                    cell.value = fundos[chaves[count]].preco_teto
                    flag+=1
                elif flag == 5:
                    cell.value = fundos[chaves[count]].pvp
                    flag+=1
                elif flag == 6:
                    cell.value = (fundos[chaves[count]].dy_percent/100)
                    flag+=1
                elif flag == 7:
                    cell.value = fundos[chaves[count]].val_patr
                    flag+=1
                elif flag == 8:
                    cell.value = fundos[chaves[count]].liquidez
                    flag+=1
                elif flag == 9:
                    cell.value = (fundos[chaves[count]].vacancia/100)
                    flag+=1
                elif flag == 10:
                    cell.value = fundos[chaves[count]].qtdcotis
                    flag+=1
        count+=1

# Programa Principal

# Carregando json
data = Data.carregar()
siglas = data.siglas
lista_fundos = {
    codigo: FundoImobiliario.from_dict(info)
    for codigo, info in data.fundos.items()
}

if data.ultima_atualizacao:
    print(f"Última atualização: {data.ultima_atualizacao}")
else:
    print("Nenhuma atualização anterior encontrada.")

fundos_filtrados = {}

sair = 0
while sair != 1:
    opcao = str(input("\n####- MENU -####\n1 - Buscar siglas\n2 - Atualizar Informações\n3 - Filtrar e atualizar tabela.\n4 - Acessar informações de um fundo\n5 - Atualizar apenas fundos filtrados.\n0 - Sair\n\nO que você deseja fazer?: "))
    if opcao == "0":
        data.siglas = siglas
        data.fundos = {k: v.to_dict() for k, v in lista_fundos.items()}
        data.ultima_atualizacao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        data.salvar()
        print(f"\nDados salvos em data.json.")
        sair = 1

    elif opcao == "1":
        siglas = obter_fiis()

    elif opcao == "2":
        for sigla in siglas:
            print(f"\nAtualizando {sigla}...")
            fundo = obter_infos(sigla)
            if fundo:
                lista_fundos[sigla] = fundo
        print("\nAtualização Completa.")

    elif opcao == "3":
        plano = openpyxl.load_workbook("Filtrados.xlsx")
        pagina = plano['Fundos']
        limpar_tabela(pagina)
        for fundo in lista_fundos.values():
            if (
                fundo.liquidez is not None and fundo.liquidez > 500_000.0 and
                fundo.pvp is not None and 0.7 <= fundo.pvp <= 1.06 and
                fundo.dy_percent is not None and 9.0 <= fundo.dy_percent <= 20.0 and
                fundo.val_patr is not None and fundo.val_patr >= 300_000_000
            ):
                fundos_filtrados[fundo.codigo] = fundo

        preencher_tabela(fundos_filtrados,pagina)
        plano.save("Filtrados.xlsx")
        print(f"\n{len(fundos_filtrados)} - Fundos Filtrados")

    elif opcao == "4":
        busca = str(input("Qual fundo você deseja verificar?: ")).upper()
        if busca in lista_fundos:
            fundo = lista_fundos[busca]
            print(f"\nCódigo: {fundo.codigo}")
            print(f"DY 12 Meses (R$): R${fundo.dy_12m:.2f}")
            print(f"DY Percentual: {fundo.dy_percent:.2f}%")
            print(f"Preço Atual: R${fundo.preco_atual:.2f}")
            print(f"Segmento: {fundo.segmento}")
            print(f"Tipo: {fundo.tipo}")
            print(f"Valor Patrimonial: R${fundo.val_patr}")
            print(f"Vacância: {fundo.vacancia:.2f}%")
            print(f"Nº de Cotistas: {fundo.qtdcotis}")
            print(f"Nº de Cotas Emitidas: {fundo.qtdcotas}")
            print(f"CNPJ: {fundo.cnpj}")
            print(f"P/VP: {fundo.pvp:.2f}")
            print(f"Liquidez Média Diária: R${fundo.liquidez:.2f}")
            print(f"Preço Teto 12% de rentabilidade: R${fundo.preco_teto:.2f}")
        else:
            print("Fundo não encontrado.")

    elif opcao == "5":
        print("\nAtualizando...")
        for codigo in fundos_filtrados.keys():
            fundo = obter_infos(codigo)
            if fundo:
                lista_fundos[codigo] = fundo
        print("\nAtualização de fundos filtrados completa.")

    else:
        print("\nDigite uma opção válida!")
